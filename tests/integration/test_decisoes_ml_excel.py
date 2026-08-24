"""Integração: 9ª aba Decisões de ML sem perder registros."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from gerar_relatorio import gerar_excel
from src.item_processor import REVISAO_ML_OFFLINE, DecisaoML
from src.operational_indicators import calcular_indicadores
from src.validacao_aula22 import (
    CLASSIFICACAO_AMBIGUO,
    CLASSIFICACAO_DIVERGENCIA,
    CLASSIFICACAO_ERRO,
    CLASSIFICACAO_VALIDO,
    RegistroValidado,
)


def _registro(classificacao: str, regra: str, lote: str) -> RegistroValidado:
    return RegistroValidado(
        lote_id=lote,
        produto="TV55-4K-B",
        linha="L1",
        turno="A",
        status_original="EM AJUSTE",
        status_normalizado="EM AJUSTE",
        responsavel="Ana",
        data="15/06/2026",
        observacao="",
        data_referencia="15/06/2026",
        classificacao=classificacao,
        regra=regra,
        mensagem="teste",
        regra_aplicada=regra,
    )


@pytest.mark.integration
def test_nona_aba_decisoes_ml_sem_perda(tmp_path: Path) -> None:
    validados = [
        _registro(CLASSIFICACAO_VALIDO, "RN08", "L1"),
        _registro(CLASSIFICACAO_AMBIGUO, "RN09", "L2"),
        _registro(CLASSIFICACAO_AMBIGUO, "RN09", "L3"),
        _registro(CLASSIFICACAO_DIVERGENCIA, "RN05", "L4"),
        _registro(CLASSIFICACAO_ERRO, "RN12", "L5"),
    ]
    decisoes = [
        DecisaoML(
            lote_id="L2",
            classe="revisar",
            probabilidade=0.72,
            nivel_confianca="média",
            latencia_ms=10.0,
            acao="revisar",
            offline=False,
            status_raw="EM AJUSTE",
            turno="A",
        ),
        DecisaoML(
            lote_id="L3",
            classe=REVISAO_ML_OFFLINE,
            probabilidade=None,
            nivel_confianca="indisponível",
            latencia_ms=None,
            acao=REVISAO_ML_OFFLINE,
            offline=True,
            status_raw="EM AJUSTE",
            turno="A",
        ),
    ]
    caminho = tmp_path / "relatorio.xlsx"
    gerar_excel(validados, calcular_indicadores(validados), caminho, decisoes_ml=decisoes)

    wb = load_workbook(caminho)
    assert "Decisões de ML" in wb.sheetnames
    for aba in (
        "Resumo",
        "Todos",
        "Válidos",
        "Divergências",
        "Ambíguos",
        "Erros de Entrada",
        "Ranking de Regras",
        "Dicionário",
    ):
        assert aba in wb.sheetnames

    ws = wb["Decisões de ML"]
    lotes = {ws.cell(row, 1).value for row in range(2, ws.max_row + 1)}
    assert lotes == {"L2", "L3"}
    assert ws.max_row - 1 == len(decisoes)
