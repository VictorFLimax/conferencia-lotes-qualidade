"""Integração: Excel consolidado com 8 abas essenciais (Aula 24)."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from gerar_relatorio import gerar_excel, gerar_resumo_executivo
from src.operational_indicators import calcular_indicadores
from src.validacao_aula22 import (
    CLASSIFICACAO_AMBIGUO,
    CLASSIFICACAO_DIVERGENCIA,
    CLASSIFICACAO_ERRO,
    CLASSIFICACAO_VALIDO,
    RegistroValidado,
)

ABAS_ESSENCIAIS = [
    "Resumo",
    "Todos",
    "Válidos",
    "Divergências",
    "Ambíguos",
    "Erros de Entrada",
    "Ranking de Regras",
    "Dicionário",
]


def _registro(classificacao: str, regra: str, lote: str, dia: str = "15/06/2026") -> RegistroValidado:
    return RegistroValidado(
        lote_id=lote,
        produto="TV55-4K-B",
        linha="L1",
        turno="A",
        status_original="APROVADO",
        status_normalizado="APROVADO",
        responsavel="Ana",
        data=dia,
        observacao="ok",
        data_referencia=dia,
        classificacao=classificacao,
        regra=regra,
        mensagem="teste",
        regra_aplicada=regra,
    )


@pytest.mark.integration
def test_relatorio_consolidado_oito_abas(tmp_path: Path) -> None:
    validados = [
        _registro(CLASSIFICACAO_VALIDO, "RN08", "L1"),
        _registro(CLASSIFICACAO_DIVERGENCIA, "RN11", "L2"),
        _registro(CLASSIFICACAO_AMBIGUO, "RN09", "L3"),
        _registro(CLASSIFICACAO_ERRO, "RN12", "L4"),
        _registro(CLASSIFICACAO_VALIDO, "RN08", "L5", "16/06/2026"),
    ]
    indicadores = calcular_indicadores(validados)

    caminho_xlsx = tmp_path / "relatorio_conferencia_lotes.xlsx"
    caminho_md = tmp_path / "resumo_executivo.md"
    gerar_excel(validados, indicadores, caminho_xlsx)
    gerar_resumo_executivo(indicadores, caminho_md)

    assert caminho_xlsx.exists()
    assert caminho_xlsx.stat().st_size > 0

    wb = load_workbook(caminho_xlsx)
    for aba in ABAS_ESSENCIAIS:
        assert aba in wb.sheetnames, f"Aba ausente: {aba}"

    ws_resumo = wb["Resumo"]
    assert len(ws_resumo._charts) >= 2
    tipos = {type(ch).__name__ for ch in ws_resumo._charts}
    assert "DoughnutChart" in tipos
    assert "LineChart" in tipos

    ranking = wb["Ranking de Regras"]
    assert ranking["B4"].value == indicadores.regra_mais_acionada_codigo
    assert ranking["D4"].value == indicadores.regra_mais_acionada_qtd

    dicionario = wb["Dicionário"]
    assert dicionario["A4"].value is not None

    texto = caminho_md.read_text(encoding="utf-8")
    assert str(indicadores.total_registros) in texto
    assert indicadores.regra_mais_acionada_codigo in texto
    assert f"{indicadores.ganho_tempo_minutos:.1f}" in texto
