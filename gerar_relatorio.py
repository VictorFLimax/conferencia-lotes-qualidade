"""
gerar_relatorio.py — Aula 22 (AX Academy)

Lê a planilha de inspeção de 10 dias, valida cada registro (RN01–RN12) e gera
o Excel `relatorio_conferencia_lotes.xlsx` com dashboard nativo (openpyxl.chart).

Uso:
  python gerar_relatorio.py

Variáveis de ambiente (.env):
  INPUT_FILE   — planilha de entrada (default: dados_entrada/inspecao_lotes_10dias_sem gabarito.xlsx)
  OUTPUT_FILE  — Excel de saída (default: relatorio_conferencia_lotes.xlsx na raiz)
  LOG_FILE     — log texto (default: logs/relatorio_aula22.log)
"""
from __future__ import annotations

import os
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.item_processor import DecisaoML, processar_ambiguos_com_ml
from src.operational_indicators import (
    META_QUALIDADE_ENTRADA_PCT,
    META_RETRABALHO_PCT,
    META_REVISAO_HUMANA_PCT,
    OperationalIndicators,
    calcular_indicadores,
)
from src.validacao_aula22 import (
    CLASSIFICACAO_AMBIGUO,
    CLASSIFICACAO_DIVERGENCIA,
    CLASSIFICACAO_ERRO,
    CLASSIFICACAO_VALIDO,
    RegistroValidado,
    validar_registro,
)

RAIZ = Path(__file__).resolve().parent
PADRAO_ABA_DIARIA = re.compile(r"^Insp_(\d{2})_(\d{2})_2026$")
COLUNAS_DIARIAS = [
    "lote_id",
    "produto",
    "linha",
    "turno",
    "status",
    "responsavel",
    "data",
    "observacao",
]


def _resolver(caminho: Path) -> Path:
    if caminho.is_absolute():
        return caminho
    return RAIZ / caminho


def _carregar_caminhos() -> tuple[Path, Path, Path, Path, Path]:
    """Resolve caminhos de entrada e saídas a partir do .env."""
    load_dotenv(RAIZ / ".env", override=False)

    entrada = _resolver(
        Path(
            os.getenv(
                "INPUT_FILE",
                "dados_entrada/inspecao_lotes_10dias_sem gabarito.xlsx",
            )
        )
    )
    saida = _resolver(Path(os.getenv("OUTPUT_FILE", "relatorio_conferencia_lotes.xlsx")))
    log = _resolver(Path(os.getenv("LOG_FILE", "logs/relatorio_aula22.log")))
    resumo_md = _resolver(Path(os.getenv("RESUMO_MD", "resumo_executivo.md")))
    json_execucao = _resolver(
        Path(os.getenv("RESUMO_JSON", "logs/resumo_execucao.json"))
    )
    return entrada, saida, log, resumo_md, json_execucao


def _data_referencia_da_aba(nome_aba: str) -> str:
    """Extrai DD/MM/AAAA a partir de Insp_DD_MM_2026."""
    m = PADRAO_ABA_DIARIA.match(nome_aba)
    if not m:
        raise ValueError(f"Nome de aba diária inválido: {nome_aba}")
    dia, mes = m.group(1), m.group(2)
    return f"{dia}/{mes}/2026"


def carregar_base_referencia(caminho: Path) -> set[str]:
    """Lê a aba Base_Referencia (cabeçalho=linha 2, dados a partir da linha 3)."""
    df = pd.read_excel(
        caminho,
        sheet_name="Base_Referencia",
        header=1,  # linha 2 do Excel (0-indexed)
        dtype=str,
    )
    if "lote_id" not in df.columns:
        raise ValueError("Base_Referencia sem coluna lote_id.")

    lotes: set[str] = set()
    for valor in df["lote_id"].tolist():
        if valor is None or (isinstance(valor, float) and pd.isna(valor)):
            continue
        texto = str(valor).strip()
        if not texto or texto.lower() == "nan":
            continue
        # Ignora rodapé / avisos (ex.: linhas sem padrão de lote)
        if texto.upper().startswith("LG-"):
            lotes.add(texto)
    return lotes


def listar_abas_diarias(caminho: Path) -> list[str]:
    """Descobre abas Insp_DD_MM_2026 via regex (sem hardcode da lista)."""
    xl = pd.ExcelFile(caminho)
    abas = [nome for nome in xl.sheet_names if PADRAO_ABA_DIARIA.match(nome)]
    if not abas:
        raise ValueError("Nenhuma aba diária Insp_DD_MM_2026 encontrada.")
    return abas


def carregar_registros_dia(caminho: Path, nome_aba: str) -> list[dict]:
    """Lê uma aba diária: título=l1, metadados=l2, cabeçalho=l3, dados a partir da l4."""
    df = pd.read_excel(
        caminho,
        sheet_name=nome_aba,
        header=2,  # linha 3 do Excel
        dtype=str,
    )
    # Garante colunas esperadas
    for col in COLUNAS_DIARIAS:
        if col not in df.columns:
            df[col] = None

    registros: list[dict] = []
    for _, row in df.iterrows():
        bruto = {col: row.get(col) for col in COLUNAS_DIARIAS}
        # Descarta linhas totalmente vazias e o rodapé "Total de registros"
        valores = []
        for v in bruto.values():
            if v is None or (isinstance(v, float) and pd.isna(v)):
                valores.append("")
            else:
                valores.append(str(v).strip())
        if not any(valores):
            continue
        lote = valores[0]
        if lote.lower().startswith("total de registros"):
            continue
        registros.append(bruto)
    return registros


def processar(caminho_entrada: Path) -> list[RegistroValidado]:
    """Lê abas, deduplica por Counter por dia e valida cada linha."""
    lotes_ref = carregar_base_referencia(caminho_entrada)
    abas = listar_abas_diarias(caminho_entrada)

    validados: list[RegistroValidado] = []
    for aba in abas:
        data_ref = _data_referencia_da_aba(aba)
        registros = carregar_registros_dia(caminho_entrada, aba)

        # Contador de ocorrências do lote_id no DIA (antes de validar).
        # Lote vazio não entra na deduplicação: RN01 tem precedência sobre RN11.
        contagem: Counter[str] = Counter()
        ocorrencias: list[int] = []
        for reg in registros:
            if chave_nao_vazia(reg):
                chave = str(reg.get("lote_id")).strip()
                contagem[chave] += 1
                ocorrencias.append(contagem[chave])
            else:
                ocorrencias.append(1)

        for reg, ocorrencia in zip(registros, ocorrencias):
            resultado = validar_registro(
                reg,
                lotes_referencia=lotes_ref,
                data_referencia=data_ref,
                ocorrencia_no_dia=ocorrencia,
            )
            validados.append(resultado)

    return validados


def chave_nao_vazia(reg: dict) -> bool:
    lote = reg.get("lote_id")
    if lote is None or (isinstance(lote, float) and pd.isna(lote)):
        return False
    return str(lote).strip() != ""


def _df_de(validados: list[RegistroValidado]) -> pd.DataFrame:
    if not validados:
        return pd.DataFrame()
    return pd.DataFrame([r.to_dict() for r in validados])


def _validar_soma_abas(indicadores: OperationalIndicators) -> None:
    """Aceite: soma das 4 classificações == total processado."""
    soma = (
        indicadores.validos_qtd
        + indicadores.divergencias_qtd
        + indicadores.ambiguos_qtd
        + indicadores.erros_qtd
    )
    if soma != indicadores.total_registros:
        raise RuntimeError(
            f"Inconsistência de classificação: "
            f"Válidos({indicadores.validos_qtd}) + "
            f"Divergências({indicadores.divergencias_qtd}) + "
            f"Ambíguos({indicadores.ambiguos_qtd}) + "
            f"Erros({indicadores.erros_qtd}) = {soma}, "
            f"mas o total processado é {indicadores.total_registros}."
        )


def _estilo_cabecalho(ws, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _escrever_dataframe(ws, df: pd.DataFrame) -> None:
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    if df.shape[1]:
        _estilo_cabecalho(ws, df.shape[1])
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col[:50]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


def _borda_fina() -> Border:
    return Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )


def _pintar_cabecalho_linha(ws, row: int, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    fino = _borda_fina()
    for col in range(1, ncols + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.border = fino


def _montar_resumo(
    wb: Workbook,
    indicadores: OperationalIndicators,
    validados: list[RegistroValidado],
) -> None:
    """Aba Resumo: 10 indicadores + DoughnutChart + LineChart (nativos)."""
    ws = wb.create_sheet("Resumo", 0)
    fino = _borda_fina()
    titulo = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    numero = Font(name="Calibri", size=14, bold=True)

    ws["A1"] = "Conferência de Lotes — Dashboard executivo"
    ws["A1"].font = titulo
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    ws["A4"] = "Os 10 indicadores operacionais"
    ws["A4"].font = Font(name="Calibri", size=12, bold=True, color="1F4E79")

    for i, h in enumerate(["#", "Indicador", "Valor", "Referência visual"], start=1):
        ws.cell(5, i, h)
    _pintar_cabecalho_linha(ws, 5, 4)

    regra_txt = (
        f"{indicadores.regra_mais_acionada_codigo} — "
        f"{indicadores.regra_mais_acionada_nome} "
        f"({indicadores.regra_mais_acionada_qtd} ocorrências)"
        if indicadores.regra_mais_acionada_codigo
        else "nenhuma"
    )
    ganho_txt = (
        f"{indicadores.ganho_tempo_minutos:.1f} min "
        f"({indicadores.ganho_tempo_horas:.2f} h)"
    )
    linhas_ind = [
        ("1", "Total de registros", str(indicadores.total_registros), "—"),
        (
            "2",
            "Registros válidos",
            f"{indicadores.validos_qtd} ({indicadores.validos_pct:.1f}%)",
            "informativa",
        ),
        (
            "3",
            "Divergências",
            f"{indicadores.divergencias_qtd} ({indicadores.divergencias_pct:.1f}%)",
            "informativa",
        ),
        (
            "4",
            "Ambíguos",
            f"{indicadores.ambiguos_qtd} ({indicadores.ambiguos_pct:.1f}%)",
            "informativa",
        ),
        (
            "5",
            "Erros de Entrada",
            f"{indicadores.erros_qtd} ({indicadores.erros_pct:.1f}%)",
            "informativa",
        ),
        ("6", "Regra mais acionada", regra_txt, "—"),
        (
            "7",
            "Taxa de qualidade da entrada",
            f"{indicadores.taxa_qualidade_entrada:.1f}%",
            f"> {META_QUALIDADE_ENTRADA_PCT:.0f}%",
        ),
        (
            "8",
            "Taxa de revisão humana",
            f"{indicadores.taxa_revisao_humana:.1f}%",
            f"< {META_REVISAO_HUMANA_PCT:.0f}%",
        ),
        (
            "9",
            "Taxa de retrabalho",
            f"{indicadores.taxa_retrabalho:.1f}%",
            f"< {META_RETRABALHO_PCT:.0f}%",
        ),
        ("10", "Ganho estimado de tempo", ganho_txt, "estimativa didática"),
    ]
    sinais = {
        7: indicadores.taxa_qualidade_entrada >= META_QUALIDADE_ENTRADA_PCT,
        8: indicadores.taxa_revisao_humana < META_REVISAO_HUMANA_PCT,
        9: indicadores.taxa_retrabalho < META_RETRABALHO_PCT,
    }
    for i, valores in enumerate(linhas_ind):
        row = 6 + i
        for col, valor in enumerate(valores, start=1):
            cell = ws.cell(row, col, valor)
            cell.border = fino
        if i + 1 in sinais:
            cor = "C6EFCE" if sinais[i + 1] else "FFC7CE"
            ws.cell(row, 4).fill = PatternFill("solid", fgColor=cor)

    # Distribuição (alimenta o gráfico de rosca) — valores do mesmo objeto
    ws["A18"] = "Distribuição por classificação"
    ws["A18"].font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    for i, h in enumerate(["Classificação", "Quantidade", "Percentual"], start=1):
        ws.cell(19, i, h)
    _pintar_cabecalho_linha(ws, 19, 3)

    dist = [
        (CLASSIFICACAO_VALIDO, indicadores.validos_qtd, indicadores.validos_pct, "C6EFCE"),
        (CLASSIFICACAO_DIVERGENCIA, indicadores.divergencias_qtd, indicadores.divergencias_pct, "FFC7CE"),
        (CLASSIFICACAO_AMBIGUO, indicadores.ambiguos_qtd, indicadores.ambiguos_pct, "FFEB9C"),
        (CLASSIFICACAO_ERRO, indicadores.erros_qtd, indicadores.erros_pct, "D9D9D9"),
    ]
    for i, (nome, qtd, pct, cor) in enumerate(dist):
        row = 20 + i
        ws.cell(row, 1, nome).border = fino
        ws.cell(row, 2, qtd).border = fino
        ws.cell(row, 2).font = numero
        ws.cell(row, 3, pct / 100).border = fino
        ws.cell(row, 3).number_format = "0.0%"
        fill = PatternFill("solid", fgColor=cor)
        for c in range(1, 4):
            ws.cell(row, c).fill = fill

    doughnut = DoughnutChart()
    doughnut.title = "Distribuição por classificação"
    labels = Reference(ws, min_col=1, min_row=20, max_row=23)
    data = Reference(ws, min_col=2, min_row=19, max_row=23)
    doughnut.add_data(data, titles_from_data=True)
    doughnut.set_categories(labels)
    doughnut.dataLabels = DataLabelList()
    doughnut.dataLabels.showPercent = True
    doughnut.dataLabels.showVal = False
    doughnut.dataLabels.showCatName = False
    doughnut.style = 10
    doughnut.width = 12
    doughnut.height = 8
    series = doughnut.series[0]
    for idx, hex_cor in enumerate(["548235", "C00000", "BF8F00", "7F7F7F"]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = hex_cor
        series.data_points.append(pt)
    ws.add_chart(doughnut, "F4")

    ws["A25"] = "Evolução diária (tabela auxiliar do gráfico)"
    ws["A25"].font = Font(name="Calibri", size=12, bold=True, color="1F4E79")

    por_dia: dict[str, Counter] = {}
    for r in validados:
        if r.data_referencia not in por_dia:
            por_dia[r.data_referencia] = Counter()
        por_dia[r.data_referencia][r.classificacao] += 1
        por_dia[r.data_referencia]["Total"] += 1

    def _chave_data(d: str) -> tuple[int, int, int]:
        partes = d.split("/")
        if len(partes) != 3:
            return (0, 0, 0)
        return (int(partes[2]), int(partes[1]), int(partes[0]))

    dias_ordenados = sorted(por_dia.keys(), key=_chave_data)
    for i, h in enumerate(
        ["Data", "Total", "Divergências", "Ambíguos", "Divergências + Ambíguos"],
        start=1,
    ):
        ws.cell(26, i, h)
    _pintar_cabecalho_linha(ws, 26, 5)

    for i, dia in enumerate(dias_ordenados):
        row = 27 + i
        c = por_dia[dia]
        div = c[CLASSIFICACAO_DIVERGENCIA]
        amb = c[CLASSIFICACAO_AMBIGUO]
        ws.cell(row, 1, dia).border = fino
        ws.cell(row, 2, c["Total"]).border = fino
        ws.cell(row, 3, div).border = fino
        ws.cell(row, 4, amb).border = fino
        ws.cell(row, 5, div + amb).border = fino

    ultima_linha = 26 + max(len(dias_ordenados), 1)
    linha = LineChart()
    linha.title = "Evolução dos registros (Divergências + Ambíguos)"
    linha.style = 10
    linha.y_axis.title = "Quantidade"
    linha.x_axis.title = "Dia"
    linha.width = 15
    linha.height = 8
    if dias_ordenados:
        dados_linha = Reference(ws, min_col=5, min_row=26, max_row=ultima_linha)
        cats_linha = Reference(ws, min_col=1, min_row=27, max_row=ultima_linha)
        linha.add_data(dados_linha, titles_from_data=True)
        linha.set_categories(cats_linha)
        ws.add_chart(linha, "F18")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 26


def _montar_ranking(wb: Workbook, indicadores: OperationalIndicators) -> None:
    ws = wb.create_sheet("Ranking de Regras")
    ws["A1"] = "Ranking de regras acionadas"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    for i, h in enumerate(
        ["Posição", "Código", "Regra", "Ocorrências", "Percentual do total"],
        start=1,
    ):
        ws.cell(3, i, h)
    _pintar_cabecalho_linha(ws, 3, 5)
    fino = _borda_fina()
    for i, item in enumerate(indicadores.ranking_regras, start=1):
        row = 3 + i
        ws.cell(row, 1, i).border = fino
        ws.cell(row, 2, item.codigo).border = fino
        ws.cell(row, 3, item.nome).border = fino
        ws.cell(row, 4, item.ocorrencias).border = fino
        ws.cell(row, 5, item.percentual / 100).border = fino
        ws.cell(row, 5).number_format = "0.0%"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22


def _montar_dicionario(wb: Workbook) -> None:
    ws = wb.create_sheet("Dicionário")
    ws["A1"] = "Dicionário de termos do relatório"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    termos = [
        ("Válido", "Registro aceito: identificação completa e status padronizado."),
        ("Divergência", "Registro com inconsistência de negócio (lote inexistente, duplicado no dia ou reprovado sem observação)."),
        ("Ambíguo", "Status não padronizado; precisa de revisão humana antes de seguir."),
        ("Erro de Entrada", "Campo obrigatório vazio ou data fora do formato dia/mês/ano."),
        ("Status", "Situação da inspeção já padronizada (APROVADO, REPROVADO ou PENDENTE)."),
        ("Status original", "Texto exatamente como veio da planilha diária, antes da padronização."),
        ("Data referência", "Dia da inspeção extraído do nome da aba (não confundir com a data digitada na linha)."),
        ("RN05", "Lote informado não existe na base oficial de referência."),
        ("RN08", "Registro aceito com status padronizado."),
        ("RN09", "Status desconhecido (por exemplo EM AJUSTE ou CANCELADO)."),
        ("RN10", "Lote reprovado sem observação preenchida."),
        ("RN11", "Mesmo lote aparece mais de uma vez no mesmo dia; conta a partir da segunda ocorrência."),
        ("RN12", "Data de inspeção ausente ou em formato diferente de DD/MM/AAAA."),
        ("Taxa de qualidade da entrada", "Percentual de registros que não são erro de preenchimento."),
        ("Taxa de revisão humana", "Percentual de registros ambíguos que exigem análise de uma pessoa."),
        ("Taxa de retrabalho", "Percentual de divergências sobre o total processado."),
        ("Ganho estimado de tempo", "Economia calculada com premissas didáticas (2 min manual vs 5 s automático), não é medição de produção."),
        ("Regra mais acionada", "Código da regra que mais classificou registros nesta execução."),
        ("Ranking de Regras", "Lista de regras ordenada da mais para a menos acionada, com quantidade e percentual."),
    ]
    for i, h in enumerate(["Termo", "Significado"], start=1):
        ws.cell(3, i, h)
    _pintar_cabecalho_linha(ws, 3, 2)
    fino = _borda_fina()
    for i, (termo, significado) in enumerate(termos, start=1):
        ws.cell(3 + i, 1, termo).border = fino
        ws.cell(3 + i, 2, significado).border = fino
        ws.cell(3 + i, 2).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 100
    ws.row_dimensions[1].height = 22


def _montar_decisoes_ml(wb: Workbook, decisoes: list[DecisaoML] | None) -> None:
    """9ª aba — auditoria de tudo que passou pelo classificador (sem perder registro)."""
    ws = wb.create_sheet("Decisões de ML")
    linhas = [d.to_excel_row() for d in (decisoes or [])]
    df = pd.DataFrame(linhas)
    if df.empty:
        df = pd.DataFrame(
            columns=[
                "Lote",
                "Classe prevista",
                "Probabilidade",
                "Nível de confiança",
                "Latência (ms)",
                "Ação aplicada",
                "API indisponível",
                "Status original",
                "Turno",
            ]
        )
    _escrever_dataframe(ws, df)


def gerar_excel(
    validados: list[RegistroValidado],
    indicadores: OperationalIndicators,
    caminho_saida: Path,
    linhas_log: list[str] | None = None,
    decisoes_ml: list[DecisaoML] | None = None,
) -> None:
    """Gera o Excel com as 8 abas essenciais + 9ª aba Decisões de ML (Log opcional)."""
    _validar_soma_abas(indicadores)

    df_todos = _df_de(validados)
    df_validos = _df_de([r for r in validados if r.classificacao == CLASSIFICACAO_VALIDO])
    df_div = _df_de(
        [r for r in validados if r.classificacao == CLASSIFICACAO_DIVERGENCIA]
    )
    df_amb = _df_de([r for r in validados if r.classificacao == CLASSIFICACAO_AMBIGUO])
    df_erro = _df_de([r for r in validados if r.classificacao == CLASSIFICACAO_ERRO])

    for nome, df, esperada in (
        ("Válidos", df_validos, CLASSIFICACAO_VALIDO),
        ("Divergências", df_div, CLASSIFICACAO_DIVERGENCIA),
        ("Ambíguos", df_amb, CLASSIFICACAO_AMBIGUO),
        ("Erros de Entrada", df_erro, CLASSIFICACAO_ERRO),
    ):
        if not df.empty and set(df["Classificação"].unique()) - {esperada}:
            raise RuntimeError(f"Aba '{nome}' mistura classificações.")

    wb = Workbook()
    padrao = wb.active
    wb.remove(padrao)

    _montar_resumo(wb, indicadores, validados)
    for nome, df in (
        ("Todos", df_todos),
        ("Válidos", df_validos),
        ("Divergências", df_div),
        ("Ambíguos", df_amb),
        ("Erros de Entrada", df_erro),
    ):
        ws = wb.create_sheet(nome)
        _escrever_dataframe(ws, df)

    _montar_ranking(wb, indicadores)
    _montar_dicionario(wb)
    _montar_decisoes_ml(wb, decisoes_ml)

    if linhas_log:
        ws_log = wb.create_sheet("Log")
        ws_log["A1"] = "Log de execução"
        ws_log["A1"].font = Font(bold=True, size=12, color="1F4E79")
        for i, linha in enumerate(linhas_log, start=3):
            ws_log.cell(i, 1, linha)
        ws_log.column_dimensions["A"].width = 100

    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)


def gerar_resumo_executivo(indicadores: OperationalIndicators, caminho: Path) -> None:
    """Escreve o markdown de negócio a partir do mesmo objeto de indicadores."""
    regra = indicadores.regra_mais_acionada_codigo or "nenhuma"
    nome_regra = indicadores.regra_mais_acionada_nome
    if regra == "RN08":
        destaque = (
            f"A regra mais acionada foi **{regra}** ({nome_regra}), com "
            f"**{indicadores.regra_mais_acionada_qtd}** ocorrências. "
            "Na prática, a maior parte das inspeções foi aceita automaticamente."
        )
    else:
        destaque = (
            f"A regra mais acionada foi **{regra}** ({nome_regra}), com "
            f"**{indicadores.regra_mais_acionada_qtd}** ocorrências. "
            "Esse é o principal ponto de atenção do processo nesta rodada "
            "e deve orientar o treino da equipe e a revisão da coleta."
        )

    texto = f"""# Resumo executivo da conferência de lotes

## Visão Geral

Foram conferidos {indicadores.total_registros} registros de inspeção. Cada linha recebeu uma única classificação (válido, divergência, ambíguo ou erro de entrada), permitindo enxergar qualidade da coleta, necessidade de revisão humana e retrabalho.

## Indicadores Principais

- Total de registros: **{indicadores.total_registros}**
- Válidos: **{indicadores.validos_qtd}** ({indicadores.validos_pct:.1f}%)
- Divergências: **{indicadores.divergencias_qtd}** ({indicadores.divergencias_pct:.1f}%)
- Ambíguos: **{indicadores.ambiguos_qtd}** ({indicadores.ambiguos_pct:.1f}%)
- Erros de entrada: **{indicadores.erros_qtd}** ({indicadores.erros_pct:.1f}%)
- Qualidade da entrada: **{indicadores.taxa_qualidade_entrada:.1f}%** (referência visual acima de {META_QUALIDADE_ENTRADA_PCT:.0f}%)
- Revisão humana: **{indicadores.taxa_revisao_humana:.1f}%** (referência visual abaixo de {META_REVISAO_HUMANA_PCT:.0f}%)
- Retrabalho: **{indicadores.taxa_retrabalho:.1f}%** (referência visual abaixo de {META_RETRABALHO_PCT:.0f}%)

## Destaque

{destaque}

## Ganho Estimado de Tempo

Premissas usadas (não cronometradas em produção):

- Conferência manual: **{indicadores.tempo_manual_segundos:.0f} segundos** por registro ({indicadores.tempo_manual_segundos / 60:.1f} min)
- Conferência automatizada: **{indicadores.tempo_automatizado_segundos:.0f} segundos** por registro

Ganho estimado: **{indicadores.ganho_tempo_minutos:.1f} minutos** ({indicadores.ganho_tempo_horas:.2f} horas) para os {indicadores.total_registros} registros desta execução.

## Observação

O ganho de tempo é uma **estimativa didática**, calculada com as premissas acima. Não substitui medição real de produção. Para virar indicador operacional de fato, seria necessário cronometrar o tempo por lote no chão de fábrica e no robô.
"""
    caminho.parent.mkdir(parents=True, exist_ok=True)
    caminho.write_text(texto, encoding="utf-8")


def gravar_log(caminho_log: Path, indicadores: OperationalIndicators) -> list[str]:
    """Grava log texto em logs/ e devolve as linhas para a aba Log."""
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ranking = ", ".join(
        f"{item.codigo}={item.ocorrencias}" for item in indicadores.ranking_regras
    )
    linhas = [
        f"Data/hora da execução: {agora}",
        f"Total de registros processados: {indicadores.total_registros}",
        f"Válidos: {indicadores.validos_qtd} ({indicadores.validos_pct:.1f}%)",
        f"Divergências: {indicadores.divergencias_qtd} ({indicadores.divergencias_pct:.1f}%)",
        f"Ambíguos: {indicadores.ambiguos_qtd} ({indicadores.ambiguos_pct:.1f}%)",
        f"Erros de Entrada: {indicadores.erros_qtd} ({indicadores.erros_pct:.1f}%)",
        (
            "Checagem de soma: "
            f"{indicadores.validos_qtd}+"
            f"{indicadores.divergencias_qtd}+"
            f"{indicadores.ambiguos_qtd}+"
            f"{indicadores.erros_qtd} = {indicadores.total_registros}"
        ),
        f"Regra mais acionada: {indicadores.regra_mais_acionada_codigo} ({indicadores.regra_mais_acionada_qtd})",
        f"Qualidade da entrada: {indicadores.taxa_qualidade_entrada:.1f}%",
        f"Revisão humana: {indicadores.taxa_revisao_humana:.1f}%",
        f"Retrabalho: {indicadores.taxa_retrabalho:.1f}%",
        f"Ganho estimado (s): {indicadores.ganho_tempo_segundos:.1f}",
        f"Ranking: {ranking}",
    ]
    caminho_log.parent.mkdir(parents=True, exist_ok=True)
    caminho_log.write_text("\n".join(linhas) + "\n", encoding="utf-8")
    return linhas


def gravar_json_execucao(caminho: Path, indicadores: OperationalIndicators) -> None:
    import json

    payload = {
        "total_registros": indicadores.total_registros,
        "validos_qtd": indicadores.validos_qtd,
        "validos_pct": indicadores.validos_pct,
        "divergencias_qtd": indicadores.divergencias_qtd,
        "divergencias_pct": indicadores.divergencias_pct,
        "ambiguos_qtd": indicadores.ambiguos_qtd,
        "ambiguos_pct": indicadores.ambiguos_pct,
        "erros_qtd": indicadores.erros_qtd,
        "erros_pct": indicadores.erros_pct,
        "regra_mais_acionada_codigo": indicadores.regra_mais_acionada_codigo,
        "regra_mais_acionada_nome": indicadores.regra_mais_acionada_nome,
        "regra_mais_acionada_qtd": indicadores.regra_mais_acionada_qtd,
        "taxa_qualidade_entrada": indicadores.taxa_qualidade_entrada,
        "taxa_revisao_humana": indicadores.taxa_revisao_humana,
        "taxa_retrabalho": indicadores.taxa_retrabalho,
        "ganho_tempo_segundos": indicadores.ganho_tempo_segundos,
        "tempo_manual_segundos": indicadores.tempo_manual_segundos,
        "tempo_automatizado_segundos": indicadores.tempo_automatizado_segundos,
        "ranking": [
            {
                "codigo": item.codigo,
                "nome": item.nome,
                "ocorrencias": item.ocorrencias,
                "percentual": item.percentual,
            }
            for item in indicadores.ranking_regras
        ],
    }
    caminho.parent.mkdir(parents=True, exist_ok=True)
    caminho.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    entrada, saida, log_path, resumo_md, json_path = _carregar_caminhos()

    if not entrada.exists():
        print(f"ERRO: arquivo de entrada não encontrado: {entrada}", file=sys.stderr)
        print(
            "Defina INPUT_FILE no .env ou coloque a planilha no caminho padrão.",
            file=sys.stderr,
        )
        return 1

    print(f"Entrada: {entrada}")
    print(f"Saída:   {saida}")

    validados = processar(entrada)
    decisoes_ml = processar_ambiguos_com_ml(
        validados,
        caminho_jsonl=_resolver(Path("logs/decisoes_ml.jsonl")),
    )
    indicadores = calcular_indicadores(validados)
    _validar_soma_abas(indicadores)

    linhas_log = gravar_log(log_path, indicadores)
    gerar_excel(validados, indicadores, saida, linhas_log, decisoes_ml)
    gerar_resumo_executivo(indicadores, resumo_md)
    gravar_json_execucao(json_path, indicadores)

    print("--- Resultado ---")
    print(f"Total processado: {indicadores.total_registros}")
    print(
        f"  Válido: {indicadores.validos_qtd} "
        f"({indicadores.validos_pct:.1f}%)"
    )
    print(
        f"  Divergência: {indicadores.divergencias_qtd} "
        f"({indicadores.divergencias_pct:.1f}%)"
    )
    print(
        f"  Ambíguo: {indicadores.ambiguos_qtd} "
        f"({indicadores.ambiguos_pct:.1f}%)"
    )
    print(
        f"  Erro de Entrada: {indicadores.erros_qtd} "
        f"({indicadores.erros_pct:.1f}%)"
    )
    print(
        f"Soma abas filtradas: "
        f"{indicadores.validos_qtd}+"
        f"{indicadores.divergencias_qtd}+"
        f"{indicadores.ambiguos_qtd}+"
        f"{indicadores.erros_qtd} = {indicadores.total_registros}  OK"
    )
    print(
        f"Regra mais acionada: {indicadores.regra_mais_acionada_codigo} "
        f"({indicadores.regra_mais_acionada_qtd})"
    )
    print(f"Decisões de ML: {len(decisoes_ml)} (aba 'Decisões de ML')")
    print(f"Log: {log_path}")
    print(f"Relatório: {saida}")
    print(f"Resumo executivo: {resumo_md}")
    print(f"JSON: {json_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
